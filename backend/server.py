from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timedelta
import jwt
import hashlib
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import base64
from enum import Enum


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# JWT Configuration
SECRET_KEY = "your-secret-key-here"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

security = HTTPBearer()

# Enums
class UserRole(str, Enum):
    ADMIN = "admin"
    COORDINATOR = "coordinator"
    EMPLOYEE = "employee"

class RequestStatus(str, Enum):
    PENDING = "pending"
    APPROVED = "approved"
    REJECTED = "rejected"

class DayOfWeek(str, Enum):
    MONDAY = "monday"
    TUESDAY = "tuesday"
    WEDNESDAY = "wednesday"
    THURSDAY = "thursday"
    FRIDAY = "friday"
    SATURDAY = "saturday"
    SUNDAY = "sunday"

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    full_name: str
    role: UserRole
    service: str
    is_active: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class UserCreate(BaseModel):
    username: str
    email: str
    full_name: str
    password: str
    role: UserRole
    service: str

class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class Schedule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    service: str
    monday_start: Optional[str] = None
    monday_break_start: Optional[str] = None
    monday_break_end: Optional[str] = None
    monday_end: Optional[str] = None
    tuesday_start: Optional[str] = None
    tuesday_break_start: Optional[str] = None
    tuesday_break_end: Optional[str] = None
    tuesday_end: Optional[str] = None
    wednesday_start: Optional[str] = None
    wednesday_break_start: Optional[str] = None
    wednesday_break_end: Optional[str] = None
    wednesday_end: Optional[str] = None
    thursday_start: Optional[str] = None
    thursday_break_start: Optional[str] = None
    thursday_break_end: Optional[str] = None
    thursday_end: Optional[str] = None
    friday_start: Optional[str] = None
    friday_break_start: Optional[str] = None
    friday_break_end: Optional[str] = None
    friday_end: Optional[str] = None
    saturday_start: Optional[str] = None
    saturday_break_start: Optional[str] = None
    saturday_break_end: Optional[str] = None
    saturday_end: Optional[str] = None
    sunday_start: Optional[str] = None
    sunday_break_start: Optional[str] = None
    sunday_break_end: Optional[str] = None
    sunday_end: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ScheduleRequest(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employee_id: str
    requested_date: str
    request_type: str  # "schedule_change" or "day_off"
    current_schedule: Optional[str] = None
    requested_schedule: Optional[str] = None
    reason: str
    status: RequestStatus = RequestStatus.PENDING
    coordinator_response: Optional[str] = None
    processed_by: Optional[str] = None
    processed_at: Optional[datetime] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ScheduleRequestCreate(BaseModel):
    requested_date: str
    request_type: str
    current_schedule: Optional[str] = None
    requested_schedule: Optional[str] = None
    reason: str

class ScheduleRequestResponse(BaseModel):
    request_id: str
    status: RequestStatus
    response: str

class Configuration(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    background_color: str = "#ffffff"
    updated_at: datetime = Field(default_factory=datetime.utcnow)

# Helper Functions
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password: str, hashed_password: str) -> bool:
    return hash_password(password) == hashed_password

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    credentials_exception = HTTPException(
        status_code=401,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            raise credentials_exception
    except jwt.PyJWTError:
        raise credentials_exception
    
    user = await db.users.find_one({"username": username})
    if user is None:
        raise credentials_exception
    
    return User(**user)

async def get_current_active_user(current_user: User = Depends(get_current_user)):
    if not current_user.is_active:
        raise HTTPException(status_code=400, detail="Inactive user")
    return current_user

# Authentication Routes
@api_router.post("/register", response_model=User)
async def register(user_data: UserCreate):
    # Check if user already exists
    existing_user = await db.users.find_one({"username": user_data.username})
    if existing_user:
        raise HTTPException(status_code=400, detail="Username already registered")
    
    # Hash password
    hashed_password = hash_password(user_data.password)
    
    # Create user
    user_dict = user_data.dict()
    user_dict.pop("password")
    user_dict["password_hash"] = hashed_password
    user_obj = User(**user_dict)
    
    await db.users.insert_one(user_obj.dict())
    return user_obj

@api_router.post("/login", response_model=Token)
async def login(user_data: UserLogin):
    user = await db.users.find_one({"username": user_data.username})
    if not user or not verify_password(user_data.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user["username"]}, expires_delta=access_token_expires
    )
    
    user_obj = User(**user)
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/me", response_model=User)
async def read_users_me(current_user: User = Depends(get_current_active_user)):
    return current_user

# User Management Routes
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    users = await db.users.find().to_list(1000)
    return [User(**user) for user in users]

@api_router.get("/employees", response_model=List[User])
async def get_employees(current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    employees = await db.users.find({"role": UserRole.EMPLOYEE}).to_list(1000)
    return [User(**employee) for employee in employees]

# Schedule Management Routes
@api_router.post("/schedules", response_model=Schedule)
async def create_schedule(schedule_data: Schedule, current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    await db.schedules.insert_one(schedule_data.dict())
    return schedule_data

@api_router.get("/schedules", response_model=List[Schedule])
async def get_schedules(current_user: User = Depends(get_current_active_user)):
    if current_user.role == UserRole.EMPLOYEE:
        # Employees can only see their own schedules
        schedules = await db.schedules.find({"user_id": current_user.id}).to_list(1000)
    else:
        # Coordinators and admins can see all schedules
        schedules = await db.schedules.find().to_list(1000)
    
    return [Schedule(**schedule) for schedule in schedules]

@api_router.get("/schedules/{user_id}", response_model=Schedule)
async def get_user_schedule(user_id: str, current_user: User = Depends(get_current_active_user)):
    if current_user.role == UserRole.EMPLOYEE and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    schedule = await db.schedules.find_one({"user_id": user_id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    return Schedule(**schedule)

@api_router.get("/my-schedule", response_model=Schedule)
async def get_my_schedule(current_user: User = Depends(get_current_active_user)):
    schedule = await db.schedules.find_one({"user_id": current_user.id})
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    
    return Schedule(**schedule)

@api_router.put("/schedules/{schedule_id}", response_model=Schedule)
async def update_schedule(schedule_id: str, schedule_data: Schedule, current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    await db.schedules.update_one(
        {"id": schedule_id},
        {"$set": schedule_data.dict()}
    )
    return schedule_data

# Schedule Request Routes
@api_router.post("/schedule-requests", response_model=ScheduleRequest)
async def create_schedule_request(request_data: ScheduleRequestCreate, current_user: User = Depends(get_current_active_user)):
    if current_user.role != UserRole.EMPLOYEE:
        raise HTTPException(status_code=403, detail="Only employees can create schedule requests")
    
    request_dict = request_data.dict()
    request_dict["employee_id"] = current_user.id
    request_obj = ScheduleRequest(**request_dict)
    
    await db.schedule_requests.insert_one(request_obj.dict())
    return request_obj

@api_router.get("/schedule-requests", response_model=List[ScheduleRequest])
async def get_schedule_requests(current_user: User = Depends(get_current_active_user)):
    if current_user.role == UserRole.EMPLOYEE:
        # Employees can only see their own requests
        requests = await db.schedule_requests.find({"employee_id": current_user.id}).to_list(1000)
    else:
        # Coordinators and admins can see all requests
        requests = await db.schedule_requests.find().to_list(1000)
    
    return [ScheduleRequest(**request) for request in requests]

@api_router.get("/pending-requests", response_model=List[ScheduleRequest])
async def get_pending_requests(current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    requests = await db.schedule_requests.find({"status": RequestStatus.PENDING}).to_list(1000)
    return [ScheduleRequest(**request) for request in requests]

@api_router.put("/schedule-requests/{request_id}/respond", response_model=ScheduleRequest)
async def respond_to_request(request_id: str, response_data: ScheduleRequestResponse, current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    update_data = {
        "status": response_data.status,
        "coordinator_response": response_data.response,
        "processed_by": current_user.id,
        "processed_at": datetime.utcnow()
    }
    
    await db.schedule_requests.update_one(
        {"id": request_id},
        {"$set": update_data}
    )
    
    updated_request = await db.schedule_requests.find_one({"id": request_id})
    return ScheduleRequest(**updated_request)

# Excel Import/Export Routes
@api_router.get("/download-template")
async def download_template(current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    # Create Excel template
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Horarios"
    
    # Headers - exact format requested
    headers = [
        "Nombre", "Servicio", "Desde", "Hasta",
        "Lunes INICIO JORNADA", "Lunes INICIO DESCANSO", "Lunes FIN DESCANSO", "Lunes FIN JORNADA",
        "Martes INICIO JORNADA", "Martes INICIO DESCANSO", "Martes FIN DESCANSO", "Martes FIN JORNADA",
        "miercoles INICIO JORNADA", "miercoles INICIO DESCANSO", "miercoles FIN DESCANSO", "miercoles FIN JORNADA",
        "Jueves INICIO JORNADA", "Jueves INICIO DESCANSO", "Jueves FIN DESCANSO", "Jueves FIN JORNADA",
        "Viernes INICIO JORNADA", "Viernes INICIO DESCANSO", "Viernes FIN DESCANSO", "Viernes FIN JORNADA",
        "Sábado INICIO JORNADA", "Sábado INICIO DESCANSO", "Sábado FIN DESCANSO", "Sábado FIN JORNADA",
        "Domingo INICIO JORNADA", "Domingo INICIO DESCANSO", "Domingo FIN DESCANSO", "Domingo FIN JORNADA"
    ]
    
    # Style headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Add sample data row
    sample_data = [
        "Juan Pérez", "Administración", "2024-01-01", "2024-12-31",
        "08:00", "12:00", "13:00", "17:00",  # Lunes
        "08:00", "12:00", "13:00", "17:00",  # Martes
        "08:00", "12:00", "13:00", "17:00",  # Miércoles
        "08:00", "12:00", "13:00", "17:00",  # Jueves
        "08:00", "12:00", "13:00", "17:00",  # Viernes
        "", "", "", "",  # Sábado (vacío)
        "", "", "", ""   # Domingo (vacío)
    ]
    
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = value
        cell.font = Font(italic=True, color="666666")
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_horarios.xlsx"}
    )

@api_router.post("/import-schedules")
async def import_schedules(file: UploadFile = File(...), current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    
    try:
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        imported_count = 0
        created_users = 0
        
        for _, row in df.iterrows():
            # Skip empty rows
            if pd.isna(row.get("Nombre")) or not row.get("Nombre"):
                continue
                
            employee_name = str(row["Nombre"]).strip()
            service_name = str(row["Servicio"]).strip()
            
            # Check if user exists by name
            user = await db.users.find_one({"full_name": employee_name})
            
            # If user doesn't exist, create them
            if not user:
                # Generate username from name
                username = employee_name.lower().replace(" ", "_").replace(".", "")
                # Check if username already exists
                existing_username = await db.users.find_one({"username": username})
                if existing_username:
                    username = f"{username}_{str(uuid.uuid4())[:8]}"
                
                # Create new user
                user_data = {
                    "id": str(uuid.uuid4()),
                    "username": username,
                    "email": f"{username}@empresa.com",
                    "full_name": employee_name,
                    "password_hash": hash_password("123456"),  # Default password
                    "role": UserRole.EMPLOYEE,
                    "service": service_name,
                    "is_active": True,
                    "created_at": datetime.utcnow()
                }
                
                await db.users.insert_one(user_data)
                user = user_data
                created_users += 1
            
            # Create or update schedule
            schedule_data = {
                "id": str(uuid.uuid4()),
                "user_id": user["id"],
                "service": service_name,
                "monday_start": str(row.get("Lunes INICIO JORNADA", "")) if not pd.isna(row.get("Lunes INICIO JORNADA")) else None,
                "monday_break_start": str(row.get("Lunes INICIO DESCANSO", "")) if not pd.isna(row.get("Lunes INICIO DESCANSO")) else None,
                "monday_break_end": str(row.get("Lunes FIN DESCANSO", "")) if not pd.isna(row.get("Lunes FIN DESCANSO")) else None,
                "monday_end": str(row.get("Lunes FIN JORNADA", "")) if not pd.isna(row.get("Lunes FIN JORNADA")) else None,
                "tuesday_start": str(row.get("Martes INICIO JORNADA", "")) if not pd.isna(row.get("Martes INICIO JORNADA")) else None,
                "tuesday_break_start": str(row.get("Martes INICIO DESCANSO", "")) if not pd.isna(row.get("Martes INICIO DESCANSO")) else None,
                "tuesday_break_end": str(row.get("Martes FIN DESCANSO", "")) if not pd.isna(row.get("Martes FIN DESCANSO")) else None,
                "tuesday_end": str(row.get("Martes FIN JORNADA", "")) if not pd.isna(row.get("Martes FIN JORNADA")) else None,
                "wednesday_start": str(row.get("miercoles INICIO JORNADA", "")) if not pd.isna(row.get("miercoles INICIO JORNADA")) else None,
                "wednesday_break_start": str(row.get("miercoles INICIO DESCANSO", "")) if not pd.isna(row.get("miercoles INICIO DESCANSO")) else None,
                "wednesday_break_end": str(row.get("miercoles FIN DESCANSO", "")) if not pd.isna(row.get("miercoles FIN DESCANSO")) else None,
                "wednesday_end": str(row.get("miercoles FIN JORNADA", "")) if not pd.isna(row.get("miercoles FIN JORNADA")) else None,
                "thursday_start": str(row.get("Jueves INICIO JORNADA", "")) if not pd.isna(row.get("Jueves INICIO JORNADA")) else None,
                "thursday_break_start": str(row.get("Jueves INICIO DESCANSO", "")) if not pd.isna(row.get("Jueves INICIO DESCANSO")) else None,
                "thursday_break_end": str(row.get("Jueves FIN DESCANSO", "")) if not pd.isna(row.get("Jueves FIN DESCANSO")) else None,
                "thursday_end": str(row.get("Jueves FIN JORNADA", "")) if not pd.isna(row.get("Jueves FIN JORNADA")) else None,
                "friday_start": str(row.get("Viernes INICIO JORNADA", "")) if not pd.isna(row.get("Viernes INICIO JORNADA")) else None,
                "friday_break_start": str(row.get("Viernes INICIO DESCANSO", "")) if not pd.isna(row.get("Viernes INICIO DESCANSO")) else None,
                "friday_break_end": str(row.get("Viernes FIN DESCANSO", "")) if not pd.isna(row.get("Viernes FIN DESCANSO")) else None,
                "friday_end": str(row.get("Viernes FIN JORNADA", "")) if not pd.isna(row.get("Viernes FIN JORNADA")) else None,
                "saturday_start": str(row.get("Sábado INICIO JORNADA", "")) if not pd.isna(row.get("Sábado INICIO JORNADA")) else None,
                "saturday_break_start": str(row.get("Sábado INICIO DESCANSO", "")) if not pd.isna(row.get("Sábado INICIO DESCANSO")) else None,
                "saturday_break_end": str(row.get("Sábado FIN DESCANSO", "")) if not pd.isna(row.get("Sábado FIN DESCANSO")) else None,
                "saturday_end": str(row.get("Sábado FIN JORNADA", "")) if not pd.isna(row.get("Sábado FIN JORNADA")) else None,
                "sunday_start": str(row.get("Domingo INICIO JORNADA", "")) if not pd.isna(row.get("Domingo INICIO JORNADA")) else None,
                "sunday_break_start": str(row.get("Domingo INICIO DESCANSO", "")) if not pd.isna(row.get("Domingo INICIO DESCANSO")) else None,
                "sunday_break_end": str(row.get("Domingo FIN DESCANSO", "")) if not pd.isna(row.get("Domingo FIN DESCANSO")) else None,
                "sunday_end": str(row.get("Domingo FIN JORNADA", "")) if not pd.isna(row.get("Domingo FIN JORNADA")) else None,
                "created_at": datetime.utcnow()
            }
            
            # Clean empty string values
            for key, value in schedule_data.items():
                if value == "" or value == "nan":
                    schedule_data[key] = None
            
            # Update or create schedule
            await db.schedules.update_one(
                {"user_id": user["id"]},
                {"$set": schedule_data},
                upsert=True
            )
            imported_count += 1
        
        return {
            "message": f"Procesados {imported_count} horarios. Creados {created_users} nuevos empleados.",
            "imported_schedules": imported_count,
            "created_users": created_users
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error processing file: {str(e)}")

@api_router.get("/export-schedules")
async def export_schedules(current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    # Get all schedules with user info
    schedules = await db.schedules.find().to_list(1000)
    data = []
    
    for schedule in schedules:
        user = await db.users.find_one({"id": schedule["user_id"]})
        if user:
            data.append({
                "Nombre": user["full_name"],
                "Servicio": schedule["service"],
                "Lunes INICIO JORNADA": schedule.get("monday_start"),
                "Lunes INICIO DESCANSO": schedule.get("monday_break_start"),
                "Lunes FIN DESCANSO": schedule.get("monday_break_end"),
                "Lunes FIN JORNADA": schedule.get("monday_end"),
                "Martes INICIO JORNADA": schedule.get("tuesday_start"),
                "Martes INICIO DESCANSO": schedule.get("tuesday_break_start"),
                "Martes FIN DESCANSO": schedule.get("tuesday_break_end"),
                "Martes FIN JORNADA": schedule.get("tuesday_end"),
                "Miércoles INICIO JORNADA": schedule.get("wednesday_start"),
                "Miércoles INICIO DESCANSO": schedule.get("wednesday_break_start"),
                "Miércoles FIN DESCANSO": schedule.get("wednesday_break_end"),
                "Miércoles FIN JORNADA": schedule.get("wednesday_end"),
                "Jueves INICIO JORNADA": schedule.get("thursday_start"),
                "Jueves INICIO DESCANSO": schedule.get("thursday_break_start"),
                "Jueves FIN DESCANSO": schedule.get("thursday_break_end"),
                "Jueves FIN JORNADA": schedule.get("thursday_end"),
                "Viernes INICIO JORNADA": schedule.get("friday_start"),
                "Viernes INICIO DESCANSO": schedule.get("friday_break_start"),
                "Viernes FIN DESCANSO": schedule.get("friday_break_end"),
                "Viernes FIN JORNADA": schedule.get("friday_end"),
                "Sábado INICIO JORNADA": schedule.get("saturday_start"),
                "Sábado INICIO DESCANSO": schedule.get("saturday_break_start"),
                "Sábado FIN DESCANSO": schedule.get("saturday_break_end"),
                "Sábado FIN JORNADA": schedule.get("saturday_end"),
                "Domingo INICIO JORNADA": schedule.get("sunday_start"),
                "Domingo INICIO DESCANSO": schedule.get("sunday_break_start"),
                "Domingo FIN DESCANSO": schedule.get("sunday_break_end"),
                "Domingo FIN JORNADA": schedule.get("sunday_end"),
            })
    
    df = pd.DataFrame(data)
    output = io.BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=horarios_exportados.xlsx"}
    )

# Configuration Routes
@api_router.get("/configuration", response_model=Configuration)
async def get_configuration():
    config = await db.configurations.find_one()
    if not config:
        # Create default configuration
        default_config = Configuration()
        await db.configurations.insert_one(default_config.dict())
        return default_config
    return Configuration(**config)

@api_router.put("/configuration", response_model=Configuration)
async def update_configuration(config_data: Configuration, current_user: User = Depends(get_current_active_user)):
    if current_user.role not in [UserRole.ADMIN, UserRole.COORDINATOR]:
        raise HTTPException(status_code=403, detail="Not enough permissions")
    
    config_data.updated_at = datetime.utcnow()
    await db.configurations.update_one(
        {},
        {"$set": config_data.dict()},
        upsert=True
    )
    return config_data

# Services Routes
@api_router.get("/services")
async def get_services(current_user: User = Depends(get_current_active_user)):
    users = await db.users.find().to_list(1000)
    services = list(set([user["service"] for user in users if user.get("service")]))
    return {"services": services}

# Initialize default admin user
@api_router.post("/init-admin")
async def init_admin():
    # Check if admin already exists
    admin = await db.users.find_one({"role": UserRole.ADMIN})
    if admin:
        return {"message": "Admin user already exists"}
    
    # Create default admin
    admin_data = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "email": "admin@horarios.com",
        "full_name": "Administrador",
        "password_hash": hash_password("admin123"),
        "role": UserRole.ADMIN,
        "service": "Administración",
        "is_active": True,
        "created_at": datetime.utcnow()
    }
    
    await db.users.insert_one(admin_data)
    return {"message": "Admin user created successfully", "username": "admin", "password": "admin123"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
